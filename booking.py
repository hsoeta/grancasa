from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import os
import openpyxl
import pyexcel
import datetime

##Airhostから宿泊者情報(.xls)をダウンロードするスクリプト. 基本的に編集しない.
##パスワードなどの変更があった時のみ変更
##morezouはFirefoxユーザなので、webdriberに注意.
def get_xls():
    driver = webdriver.Firefox()
    driver.get('https://cloud.airhost.co/ja/bookings')
    time.sleep(1)
    username = driver.find_element(By.NAME, ("user[username]"))
    password = driver.find_element(By.NAME, ('user[password]'))
    username.send_keys('ここにアカウント名を入れてね')
    password.send_keys('ここにパスワードを入れてね')
    login_btn = driver.find_element(By.CLASS_NAME, "bg-pink-400")
    login_btn.click()
    driver.implicitly_wait(3)
    dl_xls_file = driver.find_element(By.ID, "download_xls").click()
    driver.implicitly_wait(3)
    driver.quit()

get_xls()

##xlsファイルのダウンロード先のディレクトリ. 自分のものに更新する.
directory_path = '~/Downloads'
file_name      = 'bookings.xls'

if os.path.isfile(os.path.join(directory_path, file_name)):
    print("ダウンロード成功、次のステップへ進みます。")
    print("                              ")
    print("------------------------------")
    print("                              ")
else:
    print("ダウンロード失敗、再度取得ステップを試みます。")
    print("                              ")
    print("------------------------------")
    print("                              ")
    get_xls()

##file_name, dest_file_nameのパスは各々の環境に応じて編集. ファイル名は変えたらダメ(bookings.xls, bookings.xlsxで固定).
pyexcel.save_book_as(file_name='~/Downloads/bookings.xls', dest_file_name='~/Downloads/bookings.xlsx')
wb = openpyxl.load_workbook('~/Downloads/bookings.xlsx')
ws = wb['Worksheet1']

now      = datetime.datetime.now()
tomorrow = now + datetime.timedelta(1)
now_str  = now.strftime('%Y-%m-%d')
tmr_str  = tomorrow.strftime('%Y-%m-%d')

i = 2
while i < ws.max_row+1:
    cell  = ws.cell(row=i, column=4)
    value = cell.value
    if value != now_str:
    ##if value != tmr_str:
        ws.delete_rows(i)
    else:
        i = i+1

i = 2
while i < ws.max_row+1:
    cell  = ws.cell(row=i, column=14)
    value = cell.value
    if value == "システムキャンセル":
        ws.delete_rows(i)
    else:
        i = i+1
    
i = 2
while i < ws.max_row+1:
    cell  = ws.cell(row=i, column=14)
    value = cell.value
    if value == "ブロックされた":
        ws.delete_rows(i)
    else:
        i = i+1

i = 2
while i < ws.max_row+1:
    cell  = ws.cell(row=i, column=31)
    value = cell.value
    print(value)
    i = i+1

print("                              ")
print("------------------------------")
print("                              ")

ws.delete_cols(4, 4)
ws.delete_cols(7, 3)
ws.delete_cols(8)
ws.delete_cols(11, 2)
ws.delete_cols(12, 10)
ws.delete_cols(13, 2)
ws.delete_cols(14, 3)
ws.delete_cols(15, 2)

ws.insert_cols(5)
ws.insert_cols(5)

src  = 17
dest = 5
i = 0
for i in range(1, ws.max_row+1):
    ws.cell(row=i, column=dest).value = ws.cell(row=i, column=src).value

ws.delete_cols(17)

src  = 16
dest = 6
i = 0
for i in range(1, ws.max_row+1):
    ws.cell(row=i, column=dest).value = ws.cell(row=i, column=src).value

ws.delete_cols(16)

i = 2
while i < ws.max_row+1:
    cell  = ws.cell(row=i, column=14)
    value = cell.value
    print(value)
    i = i+1

print("                              ")
print("------------------------------")
print("                              ")

ws.column_dimensions['A'].width = 11
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 12
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 13
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 13
ws.column_dimensions['N'].width = 55
ws.column_dimensions['O'].width = 55

##パスを各々の環境に応じて編集.
wb.save('~/Downloads/{0}月{1}日宿泊者予定.xlsx' .format(now.month, now.day))
##wb.save('~/Downloads/{0}月{1}日宿泊者予定.xlsx' .format(tomorrow.month, tomorrow.day))
os.remove('~/Downloads/bookings.xls')
os.remove('~/Downloads/bookings.xlsx')
